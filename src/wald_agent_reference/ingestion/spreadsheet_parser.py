from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..core.models import StructuredTable
from ..utils import coerce_text, compact_whitespace


@dataclass
class SpreadsheetParser:
    preview_rows: int = 5

    def parse_file(self, path: Path) -> list[StructuredTable]:
        suffix = path.suffix.lower()
        if suffix in {".csv", ".tsv"}:
            separator = "\t" if suffix == ".tsv" else ","
            frame = pd.read_csv(path, sep=separator)
            return [self._table_from_dataframe(path, frame, sheet_name="Sheet1")]
        if suffix == ".xlsx":
            return self._parse_xlsx(path)
        if suffix == ".xls":
            excel = pd.ExcelFile(path)
            tables: list[StructuredTable] = []
            for sheet_name in excel.sheet_names:
                frame = excel.parse(sheet_name=sheet_name)
                tables.append(self._table_from_dataframe(path, frame, sheet_name=sheet_name))
            return tables
        raise ValueError(f"Unsupported spreadsheet file: {path}")

    def _parse_xlsx(self, path: Path) -> list[StructuredTable]:
        workbook = load_workbook(filename=path, data_only=True)
        tables: list[StructuredTable] = []
        for worksheet in workbook.worksheets:
            rows = self._worksheet_rows_with_merged_values(worksheet)
            matrix = self._trim_matrix(rows)
            if not matrix:
                continue
            header_idx = self._find_header_row(matrix)
            headers, body_start_idx, header_row_count = self._build_headers(matrix, header_idx)
            body = matrix[body_start_idx:]
            frame = pd.DataFrame(body, columns=headers)
            source_rows = [header_idx + header_row_count + row_idx + 1 for row_idx in range(len(frame))]
            end_column = chr(ord("A") + len(headers) - 1)
            tables.append(
                self._table_from_dataframe(
                    path,
                    frame,
                    sheet_name=worksheet.title,
                    source_rows=source_rows,
                    source_range=f"A{header_idx + 1}:{end_column}{header_idx + header_row_count + len(frame)}",
                    header_rows=header_row_count,
                )
            )
        return tables

    def _table_from_dataframe(
        self,
        path: Path,
        frame: pd.DataFrame,
        sheet_name: str,
        source_rows: list[int] | None = None,
        source_range: str | None = None,
        header_rows: int = 1,
    ) -> StructuredTable:
        normalized = frame.copy()
        normalized.columns = [self._normalize_header(column, idx) for idx, column in enumerate(normalized.columns)]
        normalized = normalized.dropna(how="all").reset_index(drop=True)
        if source_rows:
            normalized.insert(0, "_source_row", source_rows[: len(normalized)])
        retrieval_text = self._build_retrieval_text(path, sheet_name, normalized)
        table_id = f"{path.stem}:{sheet_name}"
        metadata = {
            "sheet_name": sheet_name,
            "source_file": path.name,
            "table_name": sheet_name,
            "columns": [column for column in normalized.columns if not str(column).startswith("_")],
            "source_range": source_range or self._sheet_range([column for column in normalized.columns if not str(column).startswith("_")], normalized["_source_row"].tolist() if "_source_row" in normalized.columns else []),
            "header_rows": header_rows,
            "row_count": len(normalized),
        }
        return StructuredTable(
            table_id=table_id,
            source_path=path,
            source_type="spreadsheet",
            dataframe=normalized,
            metadata=metadata,
            retrieval_text=retrieval_text,
        )

    def _build_retrieval_text(self, path: Path, sheet_name: str, frame: pd.DataFrame) -> str:
        visible_columns = [column for column in frame.columns if not str(column).startswith("_")]
        headers = ", ".join(str(column) for column in visible_columns)
        preview = frame[visible_columns].head(self.preview_rows).fillna("").astype(str) if visible_columns else frame.head(self.preview_rows).fillna("").astype(str)
        preview_rows = []
        for _, row in preview.iterrows():
            preview_rows.append(" | ".join(compact_whitespace(value) for value in row.tolist()))
        body = "\n".join(preview_rows)
        return compact_whitespace(
            f"Spreadsheet {path.name} sheet {sheet_name}. Columns: {headers}. Preview rows: {body}"
        )

    @staticmethod
    def _trim_matrix(rows: list[tuple[object, ...]]) -> list[list[object]]:
        matrix = [list(row) for row in rows]
        while matrix and all(coerce_text(cell) == "" for cell in matrix[-1]):
            matrix.pop()
        if not matrix:
            return []

        max_cols = max(len(row) for row in matrix)
        for row in matrix:
            row.extend([None] * (max_cols - len(row)))

        active_cols = [
            idx
            for idx in range(max_cols)
            if any(coerce_text(row[idx]) != "" for row in matrix)
        ]
        return [[row[idx] for idx in active_cols] for row in matrix]

    @staticmethod
    def _find_header_row(matrix: list[list[object]]) -> int:
        for idx, row in enumerate(matrix):
            non_empty = sum(1 for cell in row if coerce_text(cell) != "")
            distinct_values = {coerce_text(cell) for cell in row if coerce_text(cell)}
            if len(distinct_values) == 1 and idx + 1 < len(matrix):
                continue
            if non_empty >= max(1, len(row) // 2):
                return idx
        return 0

    @staticmethod
    def _normalize_header(value: object, idx: int) -> str:
        text = coerce_text(value)
        return text if text else f"column_{idx + 1}"

    def _worksheet_rows_with_merged_values(self, worksheet: Worksheet) -> list[tuple[object, ...]]:
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        for merged_range in worksheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            value = worksheet.cell(row=min_row, column=min_col).value
            for row_idx in range(min_row - 1, max_row):
                while row_idx >= len(rows):
                    rows.append([])
                row = rows[row_idx]
                while len(row) < max_col:
                    row.append(None)
                for col_idx in range(min_col - 1, max_col):
                    if row[col_idx] in {None, ""}:
                        row[col_idx] = value
        return [tuple(row) for row in rows]

    def _build_headers(self, matrix: list[list[object]], header_idx: int) -> tuple[list[str], int, int]:
        primary = self._forward_fill_row(matrix[header_idx])
        header_rows = [primary]
        body_start_idx = header_idx + 1
        if header_idx + 1 < len(matrix) and self._looks_like_header_extension(matrix[header_idx + 1]):
            secondary = matrix[header_idx + 1]
            header_rows.append(secondary)
            body_start_idx += 1

        combined_headers: list[str] = []
        for col_idx in range(len(primary)):
            parts = []
            for row in header_rows:
                text = coerce_text(row[col_idx])
                if text and text not in parts:
                    parts.append(text)
            combined_headers.append(self._normalize_header(" ".join(parts).strip(), col_idx))
        return combined_headers, body_start_idx, len(header_rows)

    @staticmethod
    def _forward_fill_row(row: list[object]) -> list[object]:
        filled: list[object] = []
        last_value: object = None
        for cell in row:
            text = coerce_text(cell)
            if text:
                last_value = cell
                filled.append(cell)
            else:
                filled.append(last_value)
        return filled

    @staticmethod
    def _looks_like_header_extension(row: list[object]) -> bool:
        non_empty = [coerce_text(cell) for cell in row if coerce_text(cell)]
        if len(non_empty) < 2:
            return False
        numeric_like = 0
        for value in non_empty:
            try:
                float(value.replace(",", ""))
                numeric_like += 1
            except ValueError:
                continue
        return numeric_like < len(non_empty) / 2

    @staticmethod
    def _sheet_range(headers: list[object], source_rows: list[int]) -> str:
        if not headers or not source_rows:
            return ""
        end_column = chr(ord("A") + len(headers) - 1)
        return f"A{min(source_rows) - 1}:{end_column}{max(source_rows)}"
