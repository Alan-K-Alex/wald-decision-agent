from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from wald_agent_reference.ingestion.spreadsheet_parser import SpreadsheetParser


def test_spreadsheet_parser_preserves_sheet_and_headers(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Quarterly KPIs"
    sheet.append(["Metric", "Q1 2024", "Q2 2024"])
    sheet.append(["Revenue", 120, 138])
    file_path = tmp_path / "kpis.xlsx"
    workbook.save(file_path)

    tables = SpreadsheetParser().parse_file(file_path)

    assert len(tables) == 1
    table = tables[0]
    assert table.metadata["sheet_name"] == "Quarterly KPIs"
    visible_columns = [column for column in table.dataframe.columns if not str(column).startswith("_")]
    assert visible_columns == ["Metric", "Q1 2024", "Q2 2024"]
    assert "Revenue" in table.retrieval_text


def test_spreadsheet_parser_handles_merged_multirow_headers(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Messy Summary"
    sheet.merge_cells("A1:D1")
    sheet["A1"] = "Regional KPI Summary"
    sheet.append(["Region", "Revenue", None, "Risk"])
    sheet.append([None, "Actual", "Target", "Score"])
    sheet.append(["Europe", 80, 87, 8])
    file_path = tmp_path / "messy.xlsx"
    workbook.save(file_path)

    tables = SpreadsheetParser().parse_file(file_path)

    table = tables[0]
    visible_columns = [column for column in table.dataframe.columns if not str(column).startswith("_")]
    assert "Revenue Actual" in visible_columns
    assert "Revenue Target" in visible_columns
    assert "_source_row" in table.dataframe.columns
    assert table.metadata["source_range"]
